import base64
import csv
import io
import re
from collections.abc import Awaitable, Callable
from html import unescape
from pathlib import Path
from zipfile import BadZipFile

from docx import Document
from docx.opc.exceptions import PackageNotFoundError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pypdf import PdfReader
from pypdf.errors import PdfReadError

MAX_ATTACHMENT_BYTES = 10 * 1024 * 1024
MAX_TOTAL_BYTES = 20 * 1024 * 1024
MAX_TEXT_CHARS = 60_000
MAX_ATTACHMENTS = 5


def decode_gmail_data(value: str) -> bytes:
    return base64.urlsafe_b64decode(value + "=" * (-len(value) % 4)) if value else b""


def html_text(value: str) -> str:
    without_tags = re.sub(r"<[^>]+>", " ", value)
    return re.sub(r"\s+", " ", unescape(without_tags)).strip()


def extract_document_text(filename: str, mime_type: str, content: bytes) -> str:
    suffix = Path(filename).suffix.casefold()
    if mime_type.startswith("text/") or suffix in {".txt", ".csv", ".md", ".html", ".htm"}:
        decoded = content.decode("utf-8", errors="replace")
        if mime_type == "text/html" or suffix in {".html", ".htm"}:
            return html_text(decoded)
        if suffix == ".csv":
            return "\n".join(" | ".join(row) for row in csv.reader(io.StringIO(decoded)))
        return decoded
    if mime_type == "application/pdf" or suffix == ".pdf":
        return "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(content)).pages)
    if suffix == ".docx":
        document = Document(io.BytesIO(content))
        return "\n".join(paragraph.text for paragraph in document.paragraphs)
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"Лист: {sheet.title}")
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index >= 500:
                    break
                lines.append(" | ".join("" if value is None else str(value) for value in row))
        return "\n".join(lines)
    return ""


def _parts(payload: dict) -> list[dict]:
    result = [payload]
    for part in payload.get("parts", []):
        result.extend(_parts(part))
    return result


async def email_text_bundle(
    message: dict,
    attachment_loader: Callable[[str], Awaitable[bytes]],
) -> tuple[str, list[str], list[str]]:
    body_chunks: list[str] = []
    attachment_chunks: list[str] = []
    attachment_names: list[str] = []
    total_bytes = 0

    for part in _parts(message.get("payload", {})):
        body = part.get("body", {})
        filename = part.get("filename", "")
        mime_type = part.get("mimeType", "application/octet-stream")
        if filename:
            if len(attachment_names) >= MAX_ATTACHMENTS:
                continue
            attachment_names.append(filename)
            content = decode_gmail_data(body.get("data", ""))
            if not content and body.get("attachmentId"):
                content = await attachment_loader(body["attachmentId"])
            if len(content) > MAX_ATTACHMENT_BYTES or total_bytes + len(content) > MAX_TOTAL_BYTES:
                attachment_chunks.append(f"[{filename}: пропущен из-за ограничения размера]")
                continue
            total_bytes += len(content)
            try:
                extracted = extract_document_text(filename, mime_type, content)
            except (
                BadZipFile,
                EOFError,
                InvalidFileException,
                KeyError,
                OSError,
                PackageNotFoundError,
                PdfReadError,
                TypeError,
                ValueError,
            ):
                extracted = ""
            attachment_chunks.append(
                f"Вложение: {filename}\n{extracted or '[текст извлечь не удалось]'}"
            )
        elif mime_type in {"text/plain", "text/html"} and body.get("data"):
            decoded = decode_gmail_data(body["data"]).decode("utf-8", errors="replace")
            body_chunks.append(html_text(decoded) if mime_type == "text/html" else decoded)

    combined = "\n\n".join([*body_chunks, *attachment_chunks])[:MAX_TEXT_CHARS]
    warnings = []
    if len("\n\n".join([*body_chunks, *attachment_chunks])) > MAX_TEXT_CHARS:
        warnings.append("Текст сокращён до безопасного лимита.")
    return combined, attachment_names, warnings
